import tkinter as tk
from tkinter import filedialog, PhotoImage
import os
from tkinterdnd2 import TkinterDnD, DND_FILES
from PIL import Image, ImageTk
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
from tkinter import messagebox
# -------------------
# Defino estilos
# -------------------
# Definir lados
borde_fino = Side(border_style="thin", color="000000")  # negro
borde_grueso = Side(border_style="thick", color="000000")  # negro
#Definir color
relleno_gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
relleno_celeste = PatternFill(start_color="B8E2FF", end_color="B8E2FF", fill_type="solid")

# -------------------
# Estructura Hora
# -------------------
class Hora:
    def __init__(self, hh: int, mm: int):
        self.hh = hh  # Dos caracteres, ejemplo "07"
        self.mm = mm  # Dos caracteres, ejemplo "45"

    def __str__(self):
        # :02d significa: número entero (d), 2 dígitos (2), rellenar con ceros (0)
        return f"{int(self.hh):02d}:{int(self.mm):02d}"


# Funcion auxiliar de Hora para manejarla en minutos
def hora_a_minutos(hora_str):
    hh, mm = map(int, hora_str.split(":"))
    return hh * 60 + mm


# Funcion para verificar si llego tarde, devuelve los minutos tarde
def calcular_tiempo_tarde(hora_entrada_str):
    if not hora_entrada_str:
        return 0  # no llegó → ya está como ausente
    entrada = hora_a_minutos(hora_entrada_str)
    limite = 8 * 60 + 6  # 18:06 → 18*60 + 6
    if entrada >= limite:
        return entrada - (8 * 60)  # diferencia con 18:00
    return 0

#Funcion auxiliar obtener mes
def obtenerMes(fecha):
    meses = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]
    
    # Convertir string a datetime si corresponde
    if isinstance(fecha, str):
        fecha = datetime.strptime(fecha, "%d/%m/%Y")
    
    return meses[fecha.month - 1]




# -------------------
# Estructura Empleado
# -------------------
class Empleado:
    def __init__(self, nDeEmpleado: int, nombre: str, cargo: str):
        self.NDeEmpleado = nDeEmpleado
        self.Nombre = nombre
        self.Cargo = cargo
        self.Asistencias = 0
        self.MinutosTarde = 0
        self.Ausentes = 0


# -------------------
# Estructura Fecha
# -------------------
class Fecha:
    def __init__(self, dia, mes, anio):
        self.dia = dia
        self.mes = mes
        self.anio = anio

    def __str__(self):
        return f"{self.dia:02d}/{self.mes:02d}/{self.anio}"


# -------------------
# Estructura Presente
# -------------------
class Presente:
    def __init__(self, Fecha: Fecha, nDeEmpleado: int, hDeEntrada: Hora, hDeSalida: Hora):
        self.Fecha = Fecha
        self.NDeEmpleado = nDeEmpleado
        self.HDeEntrada = hDeEntrada
        self.HDeSalida = hDeSalida
        self.Tarde = 0

    def __str__(self):
        return f"Fecha: {self.Fecha}, N:{self.NDeEmpleado}, Entrada:{self.HDeEntrada}, Salida:{self.HDeSalida}"


# -------------------
# Estructura PresentesDelDia
# -------------------
class PresentesPorDia:
    def __init__(self, dia: str, fecha: datetime.date):
        self.Dia = dia  # Ejemplo: "Lunes"
        self.Fecha = fecha  # Ejemplo: "25/09/2025"
        self.ListaPresentes = []  # Lista de objetos de Presentes del dia

    def agregar_presente(self, Presente: Presente):
        self.ListaPresentes.append(Presente)


# --------------------------------------------------------------------------------------
# Empleados
# --------------------------------------------------------------------------------------
listaDeEmpleados = []


def descubrirEmpleados(listaDeEmpleados, NEmpleado, Nombre, Cargo):
    for empleado in listaDeEmpleados:
        if empleado.NDeEmpleado == NEmpleado:
            return
    nuevo = Empleado(NEmpleado, Nombre, Cargo)
    listaDeEmpleados.append(nuevo)


# --------------------------------------------------------------------------------------
# Presentes
# --------------------------------------------------------------------------------------
listaDePresentes = []


def guardarPresente(listaDePresentes, fecha, nEmpleado, entrada, salida):
    presente = Presente(fecha, nEmpleado, entrada, salida)
    listaDePresentes.append(presente)
    return presente


# --------------------------------------------------------------------------------------
# Procesar Presentes
# --------------------------------------------------------------------------------------
def obtener_dia_semana(fecha_str):
    """ Convierte una fecha tipo string "DD/MM/YYYY" a nombre del día en español. """
    dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
    return dias_semana[fecha.weekday()]


listaDePresentesPorDia = []


def agregarAusenteATodos(listaDeEmpleados, presente):
    for empleado in listaDeEmpleados:
        falsoPresente = Presente(presente.Fecha, empleado.NDeEmpleado, None, None)
        listaDePresentesPorDia[-1].agregar_presente(falsoPresente)
        empleado.Ausentes += 1


def descubrirFechas(listaDePresentesPorDia, presente: Presente):
    for p in listaDePresentesPorDia:
        if p.Fecha == presente.Fecha:
            return
    dia = obtener_dia_semana(presente.Fecha.strftime("%d/%m/%Y"))
    nuevo = PresentesPorDia(dia, presente.Fecha)
    listaDePresentesPorDia.append(nuevo)
    agregarAusenteATodos(listaDeEmpleados, presente)


def agregarPresente(listaDePresentesPorDia, presente: Presente, listaDeEmpleados):
    pos = 0
    while (listaDePresentesPorDia[pos].Fecha != presente.Fecha):
        pos += 1
    listaDePresentesPorDia[pos].ListaPresentes[presente.NDeEmpleado - 1] = presente
    listaDeEmpleados[(presente.NDeEmpleado) - 1].Ausentes -= 1
    listaDeEmpleados[(presente.NDeEmpleado) - 1].Asistencias += 1

def main(direccion):
    # --------------------------------------------------------------------------------------
    # Abro documentos
    # --------------------------------------------------------------------------------------
    wb = openpyxl.load_workbook(direccion)
    hojaDocumento1 = wb.active

    # --------------------------------------------------------------------------------------
    # Extraer Datos de los Empleados
    # --------------------------------------------------------------------------------------
    for fila in hojaDocumento1.iter_rows(min_row=3, min_col=1, max_col=7, values_only=True):
        if fila[0] is None:
            break

        nEmpleado = fila[0]
        nombre = fila[1]
        cargo = fila[3]

        if isinstance(fila[4], datetime):
            fecha = fila[4].date()
        else:
            fecha = datetime.strptime(str(fila[4]), "%d/%m/%Y").date()

        entrada = fila[5].strftime("%H:%M") if fila[5] else None
        salida = fila[6].strftime("%H:%M") if fila[6] else None

        descubrirEmpleados(listaDeEmpleados, nEmpleado, nombre, cargo)
        presente = guardarPresente(listaDePresentes, fecha, nEmpleado, entrada, salida)


    # --------------------------------------------------------------------------------------
    # Procesar Presentes
    # --------------------------------------------------------------------------------------
    listaDeEmpleados.sort(key=lambda emp: emp.NDeEmpleado)

    for presente in listaDePresentes:
        descubrirFechas(listaDePresentesPorDia, presente)
        agregarPresente(listaDePresentesPorDia, presente, listaDeEmpleados)

        minutos_tarde = calcular_tiempo_tarde(presente.HDeEntrada)
        if minutos_tarde > 0:
            listaDeEmpleados[presente.NDeEmpleado - 1].MinutosTarde += minutos_tarde
            presente.Tarde = minutos_tarde


    #--------------------------------------------------------------------------------------
    # Genero una hoja nueva para el reporte
    #--------------------------------------------------------------------------------------

    # Crear hoja nueva
    reporteHoja = wb.create_sheet(title="Reporte")

    # Relleno blanco
    relleno_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in reporteHoja.iter_rows(min_row=1, max_row=100, min_col=1, max_col=100):
        for cell in row:
            cell.fill = relleno_blanco
            cell.alignment = Alignment(horizontal="center", vertical="center")

    reporteHoja["A1"].alignment = Alignment(horizontal="left")


    indice_medio = len(listaDePresentesPorDia) // 2
    mes = obtenerMes(listaDePresentesPorDia[indice_medio].ListaPresentes[-1].Fecha)

    # Encabezado
    titulo= f'Reporte de {mes}'
    #Tamaño de primer fila
    tamanoDePrimerFila = (len(listaDePresentesPorDia) * 2) + 6
    posicionDePrimerFila= get_column_letter(tamanoDePrimerFila)
    #Titulo
    reporteHoja.merge_cells(f"A1:{posicionDePrimerFila}1")
    reporteHoja["A1"] =  titulo

    #Informacion fija
    reporteHoja.merge_cells(f"A2:A3")
    reporteHoja["A2"] =  "N° de Empleado"

    reporteHoja.merge_cells(f"B2:B3")
    reporteHoja["B2"] =  "Nombre"

    reporteHoja.merge_cells(f"C2:C3")
    reporteHoja["C2"] =  "Cargo"

    reporteHoja.merge_cells(f"D2:D3")
    reporteHoja["D2"] =  "Presentes"

    reporteHoja.merge_cells(f"E2:E3")
    reporteHoja["E2"] =  "Ausentes"

    reporteHoja.merge_cells(f"F2:F3")
    reporteHoja["F2"] =  "Mins. Tarde Totales"

    #Carga de empleados
    auxFila=4
    for empleado in listaDeEmpleados:
        reporteHoja[f"A{auxFila}"] =  empleado.NDeEmpleado
        reporteHoja[f"B{auxFila}"] =  empleado.Nombre
        reporteHoja[f"C{auxFila}"] =  empleado.Cargo
        reporteHoja[f"D{auxFila}"] =  empleado.Asistencias
        reporteHoja[f"E{auxFila}"] =  empleado.Ausentes
        reporteHoja[f"F{auxFila}"] =  empleado.MinutosTarde
        if auxFila%2:
            reporteHoja[f"A{auxFila}"].fill = relleno_celeste
            reporteHoja[f"B{auxFila}"].fill = relleno_celeste
            reporteHoja[f"C{auxFila}"].fill = relleno_celeste
            reporteHoja[f"D{auxFila}"].fill = relleno_celeste
            reporteHoja[f"E{auxFila}"].fill = relleno_celeste
            reporteHoja[f"F{auxFila}"].fill = relleno_celeste
        auxFila+=1

    #Carga de Dias
    auxColumna=7
    for dias in listaDePresentesPorDia:
        auxLetra= get_column_letter(auxColumna)
        auxLetraSiguiente =get_column_letter(auxColumna+1)
        #Defino el dia
        reporteHoja.merge_cells(f"{auxLetra}2:{auxLetraSiguiente}2")
        dia = f"{dias.Dia} {dias.Fecha.day:02d}/{dias.Fecha.month:02d}"
        reporteHoja[f"{auxLetra}2"] =  dia
        #Defino entrada y salida
        reporteHoja[f"{auxLetra}3"] =  "Entrada"
        reporteHoja[f"{auxLetraSiguiente}3"] =  "Salida"
        #Sumo la referencia de columna
        auxColumna+=2
        #Creo auxiliar para cargar los horarios
        auxFila=4
        #Carga de dias
        for presente in dias.ListaPresentes:
            if presente.HDeEntrada != None:
                reporteHoja[f"{auxLetra}{auxFila}"] =  presente.HDeEntrada
                reporteHoja[f"{auxLetraSiguiente}{auxFila}"] =  presente.HDeSalida
            else:
                reporteHoja[f"{auxLetra}{auxFila}"] =  "-"
                reporteHoja[f"{auxLetraSiguiente}{auxFila}"] =  "-"
            auxFila+=1

    #--------------------------------------------------------------------------------------
    # Formato del excel
    #--------------------------------------------------------------------------------------
    #Titulo
    i=0
    for row in reporteHoja.iter_rows(min_row=1, max_row=1, min_col=1, max_col=tamanoDePrimerFila):
        for celda in row:
            celda.font = Font(bold=True, name="Calibri", size=18)
            celda.border = Border(bottom=borde_grueso)
            i+=1
            if i==tamanoDePrimerFila:
                celda.border =Border(bottom=borde_grueso, right=borde_grueso)

    cantDeFilas=len(listaDeEmpleados)+3

    #Fila N° de empleado
    i=1
    for row in reporteHoja.iter_rows(min_row=2, max_row=cantDeFilas, min_col=1, max_col=1):
        for celda in row:
            celda.border = Border(right=borde_grueso, left=borde_grueso)
            i+=1
            if i==cantDeFilas:
                celda.border =Border(bottom=borde_grueso, right=borde_grueso, left=borde_grueso)
            if i==3:
                celda.border =Border(bottom=borde_fino, right=borde_grueso, left=borde_grueso)
            if i<3:
                celda.fill = relleno_gris

    #Fila Nombre
    i=1
    for row in reporteHoja.iter_rows(min_row=2, max_row=cantDeFilas, min_col=2, max_col=2):
        for celda in row:
            celda.border = Border(right=borde_grueso, left=borde_grueso)
            i+=1
            if i==cantDeFilas:
                celda.border =Border(bottom=borde_grueso, right=borde_grueso, left=borde_grueso)
            if i==3:
                celda.border =Border(bottom=borde_fino, right=borde_grueso, left=borde_grueso)
            if i<3:
                celda.fill = relleno_gris

                
    #Fila Cargo
    i=1
    for row in reporteHoja.iter_rows(min_row=2, max_row=cantDeFilas, min_col=3, max_col=3):
        for celda in row:
            celda.border = Border(right=borde_grueso, left=borde_grueso)
            i+=1
            if i==cantDeFilas:
                celda.border =Border(bottom=borde_grueso, right=borde_grueso, left=borde_grueso)
            if i==3:
                celda.border =Border(bottom=borde_fino, right=borde_grueso, left=borde_grueso)
            if i<3:
                celda.fill = relleno_gris

    #Fila Presentes
    i=1
    for row in reporteHoja.iter_rows(min_row=2, max_row=cantDeFilas, min_col=4, max_col=4):
        for celda in row:
            celda.border = Border(right=borde_grueso, left=borde_grueso)
            i+=1
            if i==cantDeFilas:
                celda.border =Border(bottom=borde_grueso, right=borde_grueso, left=borde_grueso)
            if i==3:
                celda.border =Border(bottom=borde_fino, right=borde_grueso, left=borde_grueso)
            if i<3:
                celda.fill = relleno_gris

    #Fila Ausentes
    i=1
    for row in reporteHoja.iter_rows(min_row=2, max_row=cantDeFilas, min_col=5, max_col=5):
        for celda in row:
            celda.border = Border(right=borde_grueso, left=borde_grueso)
            i+=1
            if i==cantDeFilas:
                celda.border =Border(bottom=borde_grueso, right=borde_grueso, left=borde_grueso)
            if i==3:
                celda.border =Border(bottom=borde_fino, right=borde_grueso, left=borde_grueso)
            if i<3:
                celda.fill = relleno_gris

    #Fila Mins. Tarde Totales
    i=1
    for row in reporteHoja.iter_rows(min_row=2, max_row=cantDeFilas, min_col=6, max_col=6):
        for celda in row:
            celda.border = Border(right=borde_grueso, left=borde_grueso)
            i+=1
            if i==cantDeFilas:
                celda.border =Border(bottom=borde_grueso, right=borde_grueso, left=borde_grueso)
            if i==3:
                celda.border =Border(bottom=borde_fino, right=borde_grueso, left=borde_grueso)
            if i<3:
                celda.fill = relleno_gris


    #Dias
    auxColumna=7
    for dias in listaDePresentesPorDia:
        auxLetra= get_column_letter(auxColumna)
        auxLetraSiguiente =get_column_letter(auxColumna+1)
        #Bordes de Dia
        reporteHoja[f"{auxLetra}2"].border = Border(
        left=borde_grueso,
        bottom=borde_grueso
    )
        reporteHoja[f"{auxLetraSiguiente}2"].border = Border(
        right=borde_grueso,
        bottom=borde_grueso
    )
        reporteHoja[f"{auxLetra}2"].fill = relleno_gris
        reporteHoja[f"{auxLetraSiguiente}2"].fill = relleno_gris

        #Bordes de entrada y salida
        reporteHoja[f"{auxLetra}3"].border = Border(
        left=borde_grueso,
        bottom=borde_fino
    )
        reporteHoja[f"{auxLetraSiguiente}3"].border = Border(
        left=borde_fino,
        right=borde_grueso,
        bottom=borde_fino
    )
        
        reporteHoja[f"{auxLetra}3"].fill = relleno_gris
        reporteHoja[f"{auxLetraSiguiente}3"].fill = relleno_gris

        #Sumo la referencia de columna
        auxColumna+=2
        #Creo auxiliar para cargar los horarios
        auxFila=4
        #Bordes de entrada y salida
        for presente in dias.ListaPresentes:      
            if auxFila !=(len(listaDeEmpleados)+3): 
                reporteHoja[f"{auxLetra}{auxFila}"].border = Border(
                left=borde_grueso,
                right=borde_fino
                )
                reporteHoja[f"{auxLetraSiguiente}{auxFila}"].border = Border(
                left=borde_fino,
                right=borde_grueso
                )
            else:
                reporteHoja[f"{auxLetra}{auxFila}"].border = Border(
                left=borde_grueso,
                right=borde_fino,
                bottom=borde_grueso
                )
                reporteHoja[f"{auxLetraSiguiente}{auxFila}"].border = Border(
                left=borde_fino,
                right=borde_grueso,
                bottom=borde_grueso
                )
            if auxFila%2:
                reporteHoja[f"{auxLetra}{auxFila}"].fill = relleno_celeste
                reporteHoja[f"{auxLetraSiguiente}{auxFila}"].fill = relleno_celeste

            auxFila+=1

    #Funcion para acomodar anchos de columnas
    def ajustar_columnas(hoja, extra=2):
        for col in hoja.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)  # letra de columna (A, B, C, ...)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Ajustar el ancho sumando un extra
            adjusted_width = max_length + extra
            hoja.column_dimensions[col_letter].width = adjusted_width
    ajustar_columnas(reporteHoja, extra=3)

    # Guardar archivo
    directorio, nombre_archivo = os.path.split(direccion)
    nombre, ext = os.path.splitext(nombre_archivo) 
    nuevo_nombre = f"{nombre} actualizado{ext}"
    nueva_ruta = os.path.join(directorio, nuevo_nombre)
    wb.save(nueva_ruta)



#--------------------------------------------------------------------------------------
# GUI
#--------------------------------------------------------------------------------------
root = TkinterDnD.Tk()

# -------------------- Configuración --------------------
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
WINDOW_WIDTH = int(screen_width * 0.28)
WINDOW_HEIGHT = int(screen_height * 0.66)

BG_COLOR = "#800020"
CONTAINER_BORDER_COLOR = "#800020"
BAR_BG = "#800020"
BAR_ICON_NORMAL = "#000000"
BAR_ICON_HOVER = "#ff6b6b"
BTN_COLOR = "#b22222"
BTN_HOVER = "#d41b1b"
BTN_CLICK = "#ff0000"
TITLE_FONT = ("Segoe UI", 14, "bold")
TEXT_FONT = ("Segoe UI", 11)
INTERACTIVE_COLOR = "#6ec1ff"
DROP_BG_IMAGE = "drag_and_drop.png"
DROP_BG_IMAGE_ACTIVE = "drag_and_drop.png"
IMAGE_FILE = os.path.join(os.path.dirname(__file__), "Grievas_logo_1.png")
file = None  #Variable para guardar la ruta del archivo

# -------------------- Funciones --------------------
def show_message(message, success=True, close_on_confirm=False):
    # Estilo según el tipo de mensaje
    bg_color = "#b22222" 
    border_color = "#000000"

    # Tamaño y posición
    msg_width = int(WINDOW_WIDTH * 0.85)
    msg_height = int(WINDOW_HEIGHT * 0.20)
    x_pos = (screen_width - msg_width) // 2
    y_pos = (screen_height - msg_height) // 2

    # Crear ventana
    msg_win = tk.Toplevel(root, highlightbackground=border_color, highlightthickness=2, bd=0)
    msg_win.geometry(f"{msg_width}x{msg_height}+{x_pos}+{y_pos}")
    msg_win.configure(bg=bg_color)
    msg_win.overrideredirect(True)

    # Texto
    msg_label = tk.Label(msg_win, text=message, fg="white", bg=bg_color, font=TEXT_FONT, wraplength=int(msg_width * 0.9))
    msg_label.pack(expand=True, pady=(10, 0))

    # Botón Confirmar
    def close_msg():
        msg_win.destroy()
        if close_on_confirm:
            root.destroy()  # Cierra toda la app

    btn = tk.Label(msg_win, text="Confirmar", bg="#d41b1b", fg="white", font=TEXT_FONT, cursor="hand2")
    btn.pack(side="bottom", pady=(0, 15), ipadx=10, ipady=5)

    btn.bind("<Button-1>", lambda e: close_msg())
    btn.bind("<Enter>", lambda e: btn.config(bg="#ff0000"))
    btn.bind("<Leave>", lambda e: btn.config(bg="#d41b1b"))


def on_close():
    root.destroy()

def choose_file():
    global file
    file_path = filedialog.askopenfilename(
        title="Elegir archivo",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*")],
    )
    if file_path:
        file = file_path
        print("Archivo seleccionado:", file)
        update_title_with_filename(file)

def update_title_with_filename(file_path):
    title_label.config(text="Archivo Cargado", fg="#6ec1ff")
    file_name = os.path.basename(file_path)
    file_name_label.config(text=file_name)

def confirmar_operacion():
    global file
    if not file:
        show_error("⚠️ Primero debés seleccionar un archivo Excel.")
        return

    try:
        # Ejecutar la función principal
        main(file)

        # Si todo salió bien, mostrar mensaje de éxito
        show_message("Operación completada correctamente.", success=True, close_on_confirm=True)


    except Exception as e:
        # Si ocurre un error, mostrar un mensaje de fallo
        show_message(f"❌ Ocurrió un error durante la ejecución:\n\n{e}", success=False)

        

# -------------------- Ventana principal --------------------
x_pos = (screen_width - WINDOW_WIDTH) // 2
y_pos = (screen_height - WINDOW_HEIGHT) // 2

root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}+{x_pos}+{y_pos}")
root.configure(bg=CONTAINER_BORDER_COLOR)
root.overrideredirect(True)
root.bind('<Escape>', lambda e: root.destroy())
root.wm_attributes("-transparentcolor", "grey")  # ← hace transparente todo lo que tenga fondo 'grey'

# -------------------- Contenedor principal --------------------
main_container = tk.Frame(root, bg=BG_COLOR, highlightbackground="#b22222", highlightthickness=3, bd=0)
main_container.place(relx=0, rely=0, relwidth=1, relheight=1)

# -------------------- Barra superior --------------------
BAR_HEIGHT = int(WINDOW_HEIGHT * 0.10)
barra = tk.Frame(main_container, bg=BAR_BG, height=BAR_HEIGHT)
barra.pack(fill='x', ipady=int(WINDOW_HEIGHT * 0.005))

# Botón cerrar (X)
close_size = int(BAR_HEIGHT * 0.55)
close_canvas = tk.Canvas(barra, width=close_size, height=close_size, bg=BAR_BG, highlightthickness=0)
close_canvas.pack(side='right', padx=12)
pad = 5
x1, y1 = pad, pad
x2, y2 = close_size - pad, close_size - pad
line_a = close_canvas.create_line(x1, y1, x2, y2, width=2, fill=BAR_ICON_NORMAL)
line_b = close_canvas.create_line(x1, y2, x2, y1, width=2, fill=BAR_ICON_NORMAL)

def on_close_enter(event):
    close_canvas.itemconfig(line_a, fill=BAR_ICON_HOVER)
    close_canvas.itemconfig(line_b, fill=BAR_ICON_HOVER)
def on_close_leave(event):
    close_canvas.itemconfig(line_a, fill=BAR_ICON_NORMAL)
    close_canvas.itemconfig(line_b, fill=BAR_ICON_NORMAL)
def on_close_press(event):
    close_canvas.move(line_a, 0.5, 0.5)
    close_canvas.move(line_b, 0.5, 0.5)
def on_close_release(event):
    close_canvas.move(line_a, -0.5, -0.5)
    close_canvas.move(line_b, -0.5, -0.5)
    on_close()

close_canvas.bind("<Enter>", on_close_enter)
close_canvas.bind("<Leave>", on_close_leave)
close_canvas.bind("<ButtonPress-1>", on_close_press)
close_canvas.bind("<ButtonRelease-1>", on_close_release)

# Permitir mover la ventana
def start_move(e):
    root.x_click, root.y_click = e.x_root, e.y_root
def stop_move(e):
    root.x_click, root.y_click = None, None
def do_move(e):
    try:
        dx, dy = e.x_root - root.x_click, e.y_root - root.y_click
        pos = root.geometry().split('+')
        w_h = pos[0]; cur_x, cur_y = int(pos[1]), int(pos[2])
        root.geometry(f"{w_h}+{cur_x + dx}+{cur_y + dy}")
        root.x_click, root.y_click = e.x_root, e.y_root
    except Exception:
        pass

barra.bind('<Button-1>', start_move)
barra.bind('<ButtonRelease-1>', stop_move)
barra.bind('<B1-Motion>', do_move)

# -------------------- Imagen superior --------------------
image_container_height = int(WINDOW_HEIGHT * 0.25 * 1.05)
image_container = tk.Frame(main_container, bg='black', height=image_container_height)
image_container.pack(fill='x', side='top')
image_container.pack_propagate(False)

img_label = tk.Label(image_container, bg='black')
img_label.place(relx=0.5, rely=0.5, anchor='center')

def load_main_image():
    try:
        img = Image.open(IMAGE_FILE)
    except FileNotFoundError:
        img_label.config(text=f"No se encontró {IMAGE_FILE}", fg='white', bg='black')
        return

    cont_width = int(root.winfo_width() * 0.8)
    ratio = img.height / img.width
    new_width = cont_width
    new_height = int(new_width * ratio)

    img_resized = img.resize((new_width, new_height), Image.LANCZOS)
    tkimg = ImageTk.PhotoImage(img_resized)

    img_label.config(image=tkimg)
    img_label.image = tkimg
    img_label.config(bg='black')

root.after(100, load_main_image)

# -------------------- Título --------------------
title_div = tk.Frame(main_container, bg=BG_COLOR)
title_div.place(relx=0.5, rely=0.33, anchor='n', relwidth=0.8, relheight=0.1)

title_label = tk.Label(title_div, text="Subir Archivo Excel", fg='white', bg=BG_COLOR, font=TITLE_FONT)
title_label.pack(expand=True)
# Label para mostrar el nombre del archivo cargado
file_name_label = tk.Label(title_div, text="", fg=INTERACTIVE_COLOR, bg=BG_COLOR, font=("Segoe UI", 10))
file_name_label.pack(anchor='center', pady=(5, 0))

# -------------------- Función de error de Zona Drag & Drop --------------------
def show_error(message):
    # Ventana de error centrada
    err_width = int(WINDOW_WIDTH * 0.85)
    err_height = int(WINDOW_HEIGHT * 0.20)
    x_pos = (screen_width - err_width) // 2
    y_pos = (screen_height - err_height) // 2
    
    error_win = tk.Toplevel(root,highlightbackground="#000000", highlightthickness=2, bd=0)
    error_win.geometry(f"{err_width}x{err_height}+{x_pos}+{y_pos}")
    error_win.configure(bg="#b22222")
    error_win.overrideredirect(True)  # Sin barra de título
    
    # Mensaje
    msg_label = tk.Label(error_win, text=message, fg="white", bg="#b22222", font=TEXT_FONT, wraplength=int(err_width*0.9))
    msg_label.pack(expand=True, pady=(10,0))

    # Botón confirmar
    def close_error():
        error_win.destroy()

    btn = tk.Label(error_win, text="Confirmar", bg="#d41b1b", fg="white", font=TEXT_FONT, cursor="hand2")
    btn.pack(side="bottom", pady=(0,15), ipadx=10, ipady=5)

    btn.bind("<Button-1>", lambda e: close_error())
    btn.bind("<Enter>", lambda e: btn.config(bg="#ff0000"))
    btn.bind("<Leave>", lambda e: btn.config(bg="#d41b1b"))
# -------------------- Zona Drag & Drop --------------------
drop_div = tk.Frame(main_container, bg='#1a1a1a')
drop_div.place(relx=0.5, rely=0.45, anchor='n', relwidth=0.8)
drop_div.pack_propagate(False)

# Registrar drop después de crear el frame
drop_div.drop_target_register(DND_FILES)

# Funciones DND
def on_drop(event):
    global file
    files = root.tk.splitlist(event.data)
    if files:
        dropped_file = files[0]
        # Validar extensión
        if dropped_file.lower().endswith((".xls", ".xlsx")):
            file = dropped_file
            print("Archivo válido:", file)
            update_title_with_filename(file)
        else:
            show_error("⚠️ Solo se aceptan archivos Excel (.xls, .xlsx).")
            file = None

def drop_enter(event):
    drop_div.config(highlightbackground="#6ec1ff")

def drop_leave(event):
    drop_div.config(highlightbackground="#33ff00")

# Vincular eventos
drop_div.dnd_bind('<<Drop>>', on_drop)
drop_div.dnd_bind('<<DropEnter>>', drop_enter)
drop_div.dnd_bind('<<DropLeave>>', drop_leave)

# Función para cargar la imagen y textos
def load_drop_image_and_text():
    drop_width = int(drop_div.winfo_width())
    try:
        img = Image.open(os.path.join(os.path.dirname(__file__), DROP_BG_IMAGE))
    except FileNotFoundError:
        print("⚠️ No se encontró la imagen de fondo del área de drop.")
        return

    ratio = img.height / img.width
    drop_height = int(drop_width * ratio)
    drop_div.config(height=drop_height)

    img_resized = img.resize((drop_width, drop_height), Image.LANCZOS)
    tk_img = ImageTk.PhotoImage(img_resized)

    bg_label = tk.Label(drop_div, image=tk_img, bg='#1a1a1a')
    bg_label.image = tk_img
    bg_label.place(x=0, y=0, relwidth=1, relheight=1)

    text_container = tk.Frame(drop_div, bg='#b22222')
    text_container.place(relx=0.5, rely=0.75, anchor='center')

    arrastrar_label = tk.Label(text_container, text="Arrastrar y soltar", fg='white', bg="#b22222", font=TEXT_FONT)
    arrastrar_label.pack(anchor='center')

    choose_label = tk.Label(text_container, text="Elegir un archivo", fg=INTERACTIVE_COLOR, bg='#b22222', font=TEXT_FONT, cursor="hand2")
    choose_label.pack(anchor='center', pady=(5,0))
    choose_label.bind('<Button-1>', lambda e: choose_file())
    choose_label.bind('<Enter>', lambda e: choose_label.config(font=("Segoe UI", 11, "underline")))
    choose_label.bind('<Leave>', lambda e: choose_label.config(font=("Segoe UI", 11)))

root.after(200, load_drop_image_and_text)


# -------------------- Botón confirmar --------------------
button_container = tk.Frame(main_container, bg=BG_COLOR)
button_container.place(relx=0.5, rely=0.91, anchor='center', relwidth=0.5, relheight=0.1)

# Colores del botón

confirm_btn = tk.Label(button_container,text="Confirmar",fg='white',bg=BTN_COLOR,font=TEXT_FONT,cursor="hand2", bd=0, relief='flat'
)
confirm_btn.pack(expand=True, fill='both', padx=10, pady=10)  

# Efectos hover / click
def on_hover(e):
    confirm_btn.config(bg=BTN_HOVER)

def on_leave(e):
    confirm_btn.config(bg=BTN_COLOR)

def on_press(e):
    confirm_btn.config(bg=BTN_CLICK)

def on_release(e):
    confirm_btn.config(bg=BTN_HOVER)
    confirmar_operacion()

# Vincular eventos
confirm_btn.bind("<Enter>", on_hover)
confirm_btn.bind("<Leave>", on_leave)
confirm_btn.bind("<ButtonPress-1>", on_press)
confirm_btn.bind("<ButtonRelease-1>", on_release)





# -------------------- Iniciar --------------------
if __name__ == '__main__':
    root.mainloop()
